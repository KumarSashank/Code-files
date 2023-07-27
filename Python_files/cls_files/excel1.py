import openpyxl as xl
wb=xl.load_workbook('/Users/kumarsashank/Desktop/Code files/Python files/cls_files/xl1.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
val=sheet.cell(1,1)
print(val.value)
print(cell.value)
no_of_rows =sheet.max_row
print(no_of_rows)
for i in range(2, no_of_rows+1) :
    cell_new = sheet.cell(i,3)
    if((cell_new.value)>75) :
        actual_attndanc_cell = sheet.cell(i,5)
        # actual_attndanc_cell.value = cell_new.value
        actual_attndanc_cell.value = "Present"
        print("if is executer")
    else :
        actual_attndanc_cell = sheet.cell(i, 5)
        actual_attndanc_cell. value = "Absent"
        print("else is executed")
wb.save('/Users/kumarsashank/Desktop/Code files/Python files/cls_files/xl1.xlsx')